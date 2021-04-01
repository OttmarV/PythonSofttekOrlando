"""
The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
openpyxl.load_workbook(filename) returns a Workbook object.
get_sheet_names() and get_sheet_by_name() help get Worksheet objects.
The square brackets in sheet[‘A1'] get Cell objects.
Cell objects have a "value" member variable with the content of that cell.
The cell() method also returns a Cell object from a sheet.
"""

import openpyxl
import os

os.chdir("Udemy")

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))

print(wb.get_sheet_names())
sheet = wb["Sheet1"]
print(type(sheet))

print(sheet["A1"].value)
print(sheet.cell(row=1, column=1).value)

for i in range(1,6):
    print(i, sheet.cell(row=i, column = 2).value)