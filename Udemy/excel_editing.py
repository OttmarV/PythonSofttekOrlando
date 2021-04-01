"""
You can view and modify a sheet's name with its "title" member variable.
Changing a cell's value is done using the square brackets, just like changing a value in a list or dictionary.
Changes you make to the workbook object can be saved with the save() method.
"""


import openpyxl
import os

os.chdir("Udemy")

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb["Sheet"]
print(sheet)

sheet["A1"] = 42
sheet["A2"] = "Hello"
wb.save("excel_editing.xlsx")

sheet2 = wb.create_sheet("Sheet2")
print(wb.sheetnames)

sheet2.title = "My New Sheet Name"
print(wb.sheetnames)
wb.save("excel_editing.xlsx")